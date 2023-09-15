import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font ,Alignment
import openpyxl
from openpyxl import load_workbook
import openpyxl as px

from openpyxl.worksheet.views import SheetView






def extract_value_cts_common_id(df):
    # Initialize a counter for the occurrences of 'CTS Common ID'
    count = 0

    for row in df.values:
        for col, cell_value in enumerate(row):
            if cell_value == 'CTS Common ID':
                count += 1
                if count == 2:
                    # Check if there is a value to the right
                    if col + 1 < len(row):
                        value = row[col + 1]
                        if pd.notna(value):
                            return value

    # If the second occurrence of 'CTS Common ID' is not found, return None
    return None







def extract_value(df, variable, column_offset):
    # Check if the variable is in the DataFrame
    if variable in df.values:
        # Find the row and column indices of the variable
        rows, cols = divmod((df == variable).values.argmax(), len(df.columns))

        # Get the value from the cell with the specified column offset
        value = df.iloc[rows, cols + column_offset]

        # Check if the value is not NaN before returning it
        if pd.notna(value):
            return value
    return None

def set_first_row_style(worksheet):
    # Create a style object for making the first row light green
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    # Apply the style to the first row
    for cell in worksheet["1"]:
        cell.fill = light_green_fill
import sys

def search_excel_for_variables(file_path, variables):
    try:
        # Read the Excel file into a pandas DataFrame
        xls = pd.ExcelFile(file_path)

        variable_values = {var: [] for var in variables}  # Initialize an empty list for each variable

        for sheet_name in xls.sheet_names:
            # Read the sheet into a DataFrame
            df = pd.read_excel(xls, sheet_name)

            for variable in variables:
                if variable == "ANTENNA MAKE - MODEL":
                    # For "ANTENNA MAKE - MODEL", use column_offset=2 (third column to the right)
                    value = extract_value(df, variable, column_offset=2)
                    
                elif variable == "RRH - WCS band (QTY/MODEL)":
                    # For "RRH - WCS band (QTY/MODEL)", use column_offset=4 (fourth column to the right)
                    value = extract_value(df, variable, column_offset=4) 
                    
                elif variable == "STRUCTURE TYPE:":
                    # For "CTS Common ID", use the custom function extract_value_cts_common_id
                    value = extract_value(df, variable, column_offset=1)   
                    
                elif variable == "CSS - SECONDARY FUNCTION ID:":
                    # For "CTS Common ID", use the custom function extract_value_cts_common_id
                    value = extract_value(df, variable, column_offset=3)    
                
                elif variable == "CTS Common ID":
                    # For "CTS Common ID", use the custom function extract_value_cts_common_id
                    value = extract_value_cts_common_id(df)
                    
                        
                else:
                    # For other variables, use column_offset=1 (one column to the right)
                    value = extract_value(df, variable, column_offset=1)

                if value is not None:
                    variable_values[variable].append(value)

        # Find the maximum length of all lists in the variable_values dictionary
        max_length = max(len(lst) for lst in variable_values.values())

        # Pad all lists to the same length with empty strings
        for variable in variable_values:
            if len(variable_values[variable]) < max_length:
                variable_values[variable].extend([''] * (max_length - len(variable_values[variable])))

        # Create a DataFrame with the fixed column names and values      
        new_file_path = 'media/Pushpendra_CIQ.xlsx'
        writer = pd.ExcelWriter(new_file_path, engine='openpyxl')

        # Create a new Excel file using openpyxl
      

        # Add the 'eNB Info' sheet to the Excel writer
       
        
        Revision_History_data = {
               "Revision History": [""]
             
        }
        new_df = pd.DataFrame(Revision_History_data)
        new_df.to_excel(writer, index=False, sheet_name='Revision History')
   
        enb_info_data = {
            "eNBId": variable_values["4-9 DIGIT SITE ID:"],
            "eNodeB Name": variable_values["CTS Common ID"],
            "Site Address":variable_values["ADDRESS:"],
            "Structure Type": variable_values["STRUCTURE TYPE:"],
            "RBS type":[''] * max_length,
            "Cabinet Controlling DUL": [''] * max_length,
            "PLMNId": ["310410"] * max_length,
            "MCC": ["310"] * max_length,
            "MNC":["410"] * max_length,
            "mncLength":["3"] * max_length,
            "numberOfSectors per DUL":[''] * max_length,
            "tac": [''] * max_length,
            "Primary DUL eNodeB Name": variable_values["CTS Common ID"],
            "Inter-DU (IDL- 2)": [''] * max_length,
            "IDLe":variable_values["Approved? (Y/N):"],
            "1st DU type": [''] * max_length,
            "1st XMU":[''] * max_length,
            "1st XMU Port 1 (on DU)": [''] * max_length,
            "1st XMU Port 2 (on DU)":[''] * max_length,
            "1st XMU Port 3 (on DU)": [''] * max_length,
            "2nd DU type":[''] * max_length,
            "2nd XMU": [''] * max_length,
            "2nd XMU Port 1": [''] * max_length,
            "2nd XMU Port 2": [''] * max_length,
            "2nd XMU Port 3": [''] * max_length,
            "IDLe BB Pair": variable_values["CSS - SECONDARY FUNCTION ID:"],
            "timezone":variable_values["REGION:"],
            "UserLabel": [''] * max_length,
        }
        new_df = pd.DataFrame(enb_info_data)
       


        data = {
            "eNBId": variable_values["4-9 DIGIT SITE ID:"],
            "EutranCellFDDId": variable_values["Soft Sector IDs"],
            "latitude": variable_values["LAT (DEC. DEG.):"],
            "latHemisphere": ["N"] * max_length,
            "longitude": variable_values["LONG (DEC. DEG.):"],
            "geoDatum": ["NAD83"] * max_length,
            "cellRange": [''] * max_length,
            "beamDirection": variable_values["AZIMUTH"],
            "Antenna Height":  [''] * max_length,
            "tmaType": ["N/A"] * max_length,
            "tmaConfiguration": ["N/A"] * max_length,
            "TMA: dlTrafficDelay": ["0"] * max_length,
            "TMA: ulTrafficDelay": ["0"] * max_length,
            "TMA: dlAttenuation": ["0"] * max_length,
            "ExternalTMA: ulGain": ["0"] * max_length,
            "noOfTxAntennas": [''] * max_length,
            "noOfRxAntennas": [''] * max_length,
            "antenna model": variable_values["ANTENNA MAKE - MODEL"],
            "Transmission Mode": ["MIMO/OSLM/TxD"] * max_length,
            "mechanicalAntennaTilt": ["0"] * max_length,
            "electricalAntennaTilt": [''] * max_length,
            "electricalAntennaTilt_2": [''] * max_length,
            "eUTRA operating band": [''] * max_length,
            "earfcnDl": [''] * max_length,
            "earfcnUl": [''] * max_length,
            "dlChannelBandwidth": [''] * max_length,
            "ulChannelBandwidth": [''] * max_length,
            "configuredOutputPower": [''] * max_length,
            "configuredMaxTxPower": ['160'] * max_length,
            "partofsectorpower": ["100"] * max_length,
            "CGI": [''] * max_length,
            "Co-Located Technology Cell": [""] * max_length,
            "USEID": [''] * max_length,
            "CSRF": ["NO"] * max_length,
            "RRU type": variable_values["RRH - WCS band (QTY/MODEL)"],
            "RBB type": [''] * max_length,
            "Antenna type": [''] * max_length,
            "user label": [''] * max_length,
            "county": variable_values["COUNTY:"],
            "cell id": [''] * max_length,
            "ci decimal": [''] * max_length
        }
        
        # Create a new DataFrame from the data dictionary
        new_df = pd.DataFrame(data)
       
        
        pci_data = {
            "EutranCellFDDId": [''] * max_length,
            "sectorId": [''] * max_length,
            "cellId": [''] * max_length,
            "PhysicalLayerCellIdGroup": [''] * max_length,
            "physicalLayerSubCellId": [''] * max_length,
            "PCI": [''] * max_length,
            "rachRootSequence": [''] * max_length,
            "Carrier": [''] * max_length,
            "DUS/XMU": [''] * max_length,
            "DU/XMU Port": [''] * max_length,
            "DU/XMU Port Expansion": [''] * max_length,
            "For all sectors identified in eUtran Parameters and eUtran NeighborRelations tabs this form must be filled out": [''] * max_length,
        }
        new_df = pd.DataFrame(pci_data)
        
        max_length = 30
 
# Define the text to be centered in the first row
        centered_text = "Transmission Line Specifications"

# Calculate the number of spaces needed to center the text
        spaces_needed = (max_length - len(centered_text)) // 2

# Create the centered row with the text surrounded by spaces
        centered_row = [" " * spaces_needed + centered_text + " " * spaces_needed]

# Create the Feeder_Misc dictionary with the centered row
        Feeder_Misc = {
     "Transmission Line Specifications": centered_row
}


        Cluster_data = {
            "eNodeB Name": [''] * max_length,
            "Cluster":  [''] * max_length,
            "Sub Network": [''] * max_length,       
# variable_values["MARKET CLUSTER:"],
            
        }
        new_df = pd.DataFrame(Cluster_data)
        
        Losses_and_Delays={
            
            "EutranCellFDD": [''] * max_length,
            "RU1 Feeder Type": ["FIBER"] * max_length,
            "RU1 Feeder Length": ["0"] * max_length,
            "RU1 DL Feeder Loss (1dB Units)=Feeder Length (ft) x Feeder Loss (dB/ft)": ["0"] * max_length,
            "RU1 Jumper Type": ["LDF4-50A"] * max_length,
            "RU1 Jumper Length": ["10"] * max_length,
            "RU1 DL Jumper Loss (1dB Units)=Jumper Length (ft) x Jumper Loss (dB/ft)": ["0.203"] * max_length,
            "RU1 Number of Connectors": ["2"] * max_length,
            "RU1 DL Connector Loss (1dB Units)= Connector Loss (sqrt (Freq in GHz) x 0.05) x No. of Connectors": ["0.0922"] * max_length,
            "RU2 Feeder Type": ["FIBER"] * max_length,
            "RU2 Feeder Length": ["0"] * max_length,
            "RU2 DL Feeder Loss (1dB Units)=Feeder Length (ft) x Feeder Loss (dB/ft)": ["0"] * max_length,
            "RU2 Jumper Type": ["LDFA-50A"] * max_length,
            "RU2 Jumper Length": ["10"] * max_length,
            "RU2 DL Jumper Loss (1dB Units)=Jumper Length (ft) x Jumper Loss (dB/ft)": ["0.203"] * max_length,
            "RU2 Number of Connectors": ["2"] * max_length,
            "RU2 DL Connector Loss (1dB Units)= Connector Loss (sqrt (Freq in GHz) x 0.05) x No. of Connectors": ["0.0922"] * max_length,
            "RU1/RU2 DL Duplexer Loss (1dB Units) From specs sheet": ["0"] * max_length,
            "RU1/RU2 DL Diplexer Loss (1dB Units) From specs sheet": ["0"] * max_length,
            "RU1/RU2 DL Other Misc.Losses (1dB Units)": ["0"] * max_length,
            "RU1/RU2 Golden Feeder TMA dlAttenuation (1dB Units)From specs sheet": ["0"] * max_length,
            "RU1 DL Feeder Delay (1nS Units)=Feeder Length (ft) / Feeder Delay (ft/nS)": ["0"] * max_length,
            "RU1 DL Jumper Delay (1nS Units)=Jumper Length (ft) / Jumper Delay (ft/nS)": ["11.547"] * max_length,
            "RU2 DL Feeder Delay (1nS Units)=Feeder Length (ft) / Feeder Delay (ft/nS)": ["0"] * max_length,
            "RU2 DL Jumper Delay (1nS Units)=Jumper Length (ft) / Jumper Delay (ft/nS)": ["11.547"] * max_length,
            "RU1/RU2 Golden Feeder TMA DL Delay (1ns Unit)From specs sheet": ["0"] * max_length,
            "RU1 Feeder Type": ["FIBER"] * max_length,
            "RU1 Feeder Length": ["0"] * max_length,
            "RU1 UL Feeder Loss (1dB Units)=Feeder Length (ft) x Feeder Loss (dB/ft)": ["0"] * max_length,
            "RU1 Jumper Type": ["LDF4-50A"] * max_length,
            "RU1 Jumper Length": ["10"] * max_length,
            "RU1 UL Jumper Loss (1dB Units)=Jumper Length (ft) x Jumper Loss (dB/ft)": ["0.203"] * max_length,
            "RU1 Number of Connectors": ["2"] * max_length,
            "RU1 UL Connector Loss (1dB Units)= Connector Loss (sqrt (Freq in GHz) x 0.05) x No. of Connectors": ["0.0922"] * max_length,
            "RU2 Feeder Length": ["0"] * max_length,
            "RU2 UL Feeder Loss (1dB Units)=Feeder Length (ft) x Feeder Loss (dB/ft)": ["0"] * max_length,
            "RU2 Jumper Type": ["LDF4-50A"] * max_length,
            "RU2 Jumper Length": ["10"] * max_length,
            "RU2 UL Jumper Loss (1dB Units)=Jumper Length (ft) x Jumper Loss (dB/ft)": ["0.203"] * max_length,
            "RU2 Number of Connectors": ["2"] * max_length,
            "RU2 UL Connector Loss (1dB Units)= Connector Loss (sqrt (Freq in GHz) x 0.05) x No. of Connectors": ["0.0922"] * max_length,
            "RU2 UL Connector Loss (1dB Units)= Connector Loss (sqrt (Freq in GHz) x 0.05) x No. of Connectors": ["0"] * max_length,
            "RU1/RU2 UL Duplexer Loss (1dB Units)From specs sheet": ["0"] * max_length,
            "RU1/RU2 UL Diplexer Loss (1dB Units)From specs Sheet": ["0"] * max_length,
            "RU1/RU2 UL Other Misc.Losses (1dB Units)": ["0"] * max_length,
            "RU1/RU2 Golden Feeder TMA ulGain (1dB Units)From specs sheet": ["0"] * max_length,
            "RU1 UL Feeder Delay (1nS Units)=Feeder Length (ft) / Feeder Delay (ft/nS)": ["0"] * max_length,
            "RU1 UL Jumper Delay (1nS Units)=Jumper Length (ft) / Jumper Delay (ft/nS)": ["11.547"] * max_length,
            "RU2 UL Feeder Delay (1nS Units)=Feeder Length (ft) / Feeder Delay (ft/nS)": ["0"] * max_length,
            "RU2 UL Jumper Delay (1nS Units)=Jumper Length (ft) / Jumper Delay (ft/nS)": ["11.547"] * max_length,
            "RU1/RU2 Golden Feeder TMA UL Delay (1ns Unit)From specs sheet": ["0"] * max_length,  
            "Delay Constant": ["20"] * max_length,
            "RFBranch: RU1 dlTrafficDelay": ["17"] * max_length,
            "RFBranch: RU1 ulTrafficDelay": ["17"] * max_length,
            "RFBranch: RU1 dlAttenuation": ["0"] * max_length,
            "RFBranch: RU1 ulAttenuation": ["0"] * max_length,
            "RFBranch: RU2 dlTrafficDelay": ["17"] * max_length,
            "RFBranch: RU2 ulTrafficDelay":["17"] * max_length,
            "RFBranch: RU2 dlAttenuation": ["0"]  * max_length,
            "RFBranch: RU2 ulAttenuation": ["0"]  * max_length,
            "RFBranch: RU3 dlTrafficDelay": ["17"] * max_length,
            "RFBranch: RU3 ulTrafficDelay": ["17"]* max_length,
            "RFBranch: RU3 dlAttenuation": ["0"] * max_length,
            "RFBranch: RU3 ulAttenuation": ["0"]  * max_length,
            "RFBranch: RU4 dlTrafficDelay":["17"] * max_length,
            "RFBranch: RU4 ulTrafficDelay": ["17"] * max_length,
            "RFBranch: RU4 dlAttenuation": ["0"] * max_length,
            "RFBranch: RU4 ulAttenuation": ["0"]  * max_length,
            
        }
        
        new_df = pd.DataFrame(Losses_and_Delays)
        
        eUtran_NeighRelations_co_sites = {
            "EutranCellFDDId": [''] * max_length,
            "EutranFreqRelationID": [''] * max_length,
            "EUTRANFreqID": [''] * max_length,
            "Neigh #1": [''] * max_length,
            "Neigh #2": [''] * max_length,
           
            
        }
        new_df = pd.DataFrame(eUtran_NeighRelations_co_sites)
        
        LTE_UMTS_UtranFreqRelation = {
            "EUtranCellFDD": [''] * max_length,
            "UtranFreqRelationId": [''] * max_length,
            "cellReselectionPriority": [''] * max_length,
            "connectedModeMobilityPrio": [''] * max_length,
            "csFallbackPrio": [''] * max_length,
            "csFallbackPrioEC": [''] * max_length,
            "ExternalUTRANFreqID": [''] * max_length,
            "uarfcn": [''] * max_length,
            "County": [''] * max_length,
           
            
        }
        new_df = pd.DataFrame(eUtran_NeighRelations_co_sites)
        
        LTE_LTE_EUtranFreqRelation = {
            "EUtranCellFDD": [''] * max_length,
            "EutranFreqRelationID": [''] * max_length,
            "cellReselectionPriority": [''] * max_length,
            "connectedModeMobilityPrio": [''] * max_length,
            "ExternalEUTRANFreqID": [''] * max_length,
            "earfcn": [''] * max_length,
            "threshXHigh": [''] * max_length,
            "threshXLow": [''] * max_length,
           
            
        }
        new_df = pd.DataFrame(eUtran_NeighRelations_co_sites)
       

        
        
       
               
        
        Revision_History_df = pd.DataFrame(Revision_History_data)
        Revision_History_df.to_excel(writer, index=False, sheet_name='Revision History')
        
        new_df.to_excel(writer, index=False, sheet_name='eNB Info')
        
        workbook = writer.book
        
      
 

# Merge and format the cell for the Transmission Line Specifications heading

       
        worksheet = workbook['Revision History']
        headers_data = {
        "Version": ["", ""],
        "Description": ["", ""],
        "Updated Date": ["", ""],
        "Updated By": ["", ""]
    }
        worksheet = workbook['Revision History']
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        for COL in worksheet.iter_rows(min_row=0, max_col=len(Revision_History_df) + 100):
         for cell in worksheet["1"]:
            cell.fill = light_green_fill
            light_green_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
        for COL in worksheet.iter_rows(min_row=0, max_col=len(Revision_History_df) + 100):
         for cell in worksheet["2"]:
            cell.fill = light_green_fill
         for cell in worksheet["3"]:
            cell.fill = light_green_fill
            light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        for COL in worksheet.iter_rows(min_row=0, max_col=len(Revision_History_df) + 100):
         for cell in worksheet["4"]:
            cell.fill = light_green_fill
        df_headers = pd.DataFrame(headers_data)

    # Write the DataFrame to the Excel file starting from the second row
        df_headers.to_excel(writer, index=False, sheet_name='Revision History', startrow=3)
        
       
        
        enb_info_df = pd.DataFrame(enb_info_data)
        enb_info_df.to_excel(writer, index=False, sheet_name='eNB Info')
               
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
                    
                    
     
        
        # Add the DataFrame to the Excel writer in 'eUtran Parameters' sheet
        new_df.to_excel(writer, index=False, sheet_name='eUtran Parameters')
        

        # Access the openpyxl workbook and worksheet objects
        workbook = writer.book
        worksheet = workbook['eNB Info']
        
        set_first_row_style(worksheet)
        
        data_df = pd.DataFrame(data)
        data_df.to_excel(writer, index=False, sheet_name='eUtran Parameters')
        
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width

        # Add the DataFrame to the Excel writer in 'eUtran Parameters' sheet
        new_df.to_excel(writer, index=False, sheet_name='PCI')
        

        # Access the openpyxl workbook and worksheet objects
        workbook = writer.book
        worksheet = workbook['eUtran Parameters']
        set_first_row_style(worksheet)
        
        pci_df = pd.DataFrame(pci_data)
        pci_df.to_excel(writer, index=False, sheet_name='PCI')
        
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        new_df.to_excel(writer, index=False, sheet_name='Feeder Misc')
        
        workbook = writer.book
        worksheet = workbook['PCI']
        set_first_row_style(worksheet)
        
        
        Feeder_Misc_df = pd.DataFrame(Feeder_Misc)
        Feeder_Misc_df.to_excel(writer, index=False, sheet_name='Feeder Misc')
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        new_df.to_excel(writer, index=False, sheet_name='Cluster')
        
        workbook = writer.book
        worksheet = workbook['Feeder Misc']
        set_first_row_style(worksheet)
        worksheet = workbook['Feeder Misc']
       
        
        
        max_length = 200

        headers_data = {
    "Manufacture": [
        "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW", "ANDREW",
        "EUPEN", "EUPEN", "EUPEN",
        "Commscope", "Commscope", "RFS", "RFS", "RFS", "RFS", "", "Commscope" ,""
    ],
    "Feeder": [
        "LDF4-50A","LDF5-50A", "LDF6-50A", "LDF7-50A","LDF12-50", "AVA5-50", "AVA7-50", "VXL5-50", "VXL6-50", "FXL 780", "FXL 1873", "EC5-50A", "EC6-50A", "EC7-50A", "CR1070 PE", "CR1873 PE", "LCF12-50J", "LCF78-50JA", "LCFS114-50JA",  "LCF158-50JA","FIBER", "CR 540 Jumper",  ""
    ],
    "Diameter": [
        "1/2\"", "7/8\"", "1 1/4\"", "1 5/8\"", "2 1/4\"", "7/8\"", "1 5/8\"", "7/8\"", "1 1/4\"", "7/8\"", "1 5/8\"",
        "7/8\"", "1 1/4\"", "1 5/8\"", "7/8\"", "1 5/8\"", "1/2\"", "7/8\"", "1 1/4\"", "1 5/8\"", "", "1/2",""
    ],
    "Loss (dB/ft)-700": [
        "0.01830", "0.01030", "0.00732", "0.00602", "0.00519", "0.00943", "0.00561", "0.01119", "0.00795", "0.00981", "0.00554", "0.00964",
        "0.00698", "0.00570",
        "0.00910", "0.00550", "0.01810", "0.00983", "0.00731", "0.00591", "0.0000", "0.0174",""
    ],
     "Loss (dB/m)-700": [
        "0.06010", "0.03370", "0.02400", "0.01974", "0.01703", "0.03093", "0.01840", "0.03670", "0.02607", "0.03219", "0.01817",
        "0.03160", "0.02290", "0.01870",
        "0.02990", "0.01820", "0.05950", "0.03230", "0.02400", "0.01940", "0.00000", "",""
    ],
    "Loss (dB/ft)-850": [
        "0.02033", "0.01149", "0.00817", "0.00674", "0.00583", "0.01047", "0.00625", "0.01244", "0.00888", "0.01090", "0.00618",
        "0.01072", "0.00776", "0.00634", "0.01015", "0.00611", "0.02013", "0.01099", "0.00816", "0.00662", "0.00000", "0.01900",""
    ],
    "Loss (dB/m)-850": [
        "0.06668", "0.03757", "0.02678", "0.02211", "0.01913", "0.03436", "0.02051", "0.04083", "0.02913", "0.03576", "0.02028",
        "0.03512", "0.02545", "0.02083", "0.03326", "0.02013", "0.06598", "0.03589", "0.02678", "0.02171", "0.00000", "",""
    ], 
    
    
     "Loss (dB/ft)-1700": [
        "", "", "", "", "", "", "", "", "", "", "",
        "", "", "", "0.01520", "", "", "", "", "", "", "0.0287",""
    ],
    "Loss (dB/m)-1700": [
        "", "", "", "", "", "", "", "", "", "", "",
        "", "", "", "", "", "", "", "", "", "", "",""
    ], 
    
    
    
    
    
    
    
    "Loss (dB/ft)-1900": [
        "0.03160", "0.01805", "0.01310", "0.01095", "0.009615", "0.016345", "0.00993", "0.019565", "0.01428", "0.017025", "0.009895", "0.01660",
        "0.01220", "0.00998","0.01630", "0.01015",
        "0.03110", "0.01720", "0.0130875", "0.01075","0", "0.03040",""
    ],
    "Loss (dB/m)-1900": [
        "0.10400", "0.05930", "0.04295", "0.035915", "0.031555", "0.05364", "0.03258", "0.06419", "0.046855", "0.055855", "0.032465", "0.05460",
        "0.04000", "0.03270", "0.05340",
        "0.03335", "0.10200", "0.05650", "0.04284", "0.03525", "0.0000","",""
    ],
    "Loss (dB/ft)-2100": [
        "0.03340", "0.01920", "0.01390", "0.01165", "0.01026", "0.01731", "0.01054", "0.02073", "0.01518", "0.01802", "0.01051", "0.0175705882352941",
        "0.01295294117647060","0.01060", "0.01730", "0.01080",
        "0.03290", "0.01830", "0.01390", "0.01140", "0.000", "0.0320",""

    ],
    "Loss (dB/m)-2100": [
        "0.1100", "0.06290", "0.04560", "0.03824", "0.03366", "0.05678", "0.03457", "0.06802", "0.04981", "0.05913", "0.03449", "0.0577294117647059",
        "0.0423764705882353","0.03460", "0.05670", "0.03540",
        "0.10800", "0.05990", "0.04550", "0.03760", "0.00","",""
    ],

      "Rel.Velocity": [
        "0.88", "0.89", "0.89", "0.88", "0.88", "0.91", "092", "0.88", "0.88", "0.88", "0.88", "0.88",
        "0.88", "0.89", "0.88",
        "0.88", "0.88", "0.89", "0.89", "0.90", "1.00","0.88",""
    ],
    "delay m/ns": [
        "0.264","0.267", "0.267", "0.264", "0.264", "0.273", "0.276", "0.264", "0.264", "0.264", "0.264", "0.267", "0.264", "0.267","0.264", "0.264", "0.264", "0.267", "0.267", "0.270", "0.300", "0.264",""
    ],
     "delay ft/ns": [

        "0.8661417336", "0.8759842533", "0.8759842533", "0.8661417336", "0.8661417336", "0.8956692927", "0.9055118124", "0.8661417336", "0.8661417336", "0.8661417336", "0.8661417336",
        "0.8759842533", "0.8661417336", "0.8759842533", "0.8661417336", "0.8661417336", "0.8661417336", "0.8759842533", "0.8759842533", "0.885826773", "0.98425197", "0.8661417336",""

    ]
}



        worksheet = workbook['Feeder Misc']
        light_green_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
        for COL in worksheet.iter_rows(min_row=0, max_col=len(Revision_History_df) + 100):
         for cell in worksheet["1"]:
            cell.fill = light_green_fill
            light_green_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
        for COL in worksheet.iter_rows(min_row=0, max_col=len(Revision_History_df) + 100):
         for cell in worksheet["2"]:
            cell.fill = light_green_fill
         for cell in worksheet["3"]:
            cell.fill = light_green_fill
         for cell in worksheet["4"]:
            cell.fill = light_green_fill
      
        df_headers = pd.DataFrame(headers_data)
        

    # Write the DataFrame to the Excel file starting from the second row
        df_headers.to_excel(writer, index=False, sheet_name='Feeder Misc', startrow=3)
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        
        
      
        Cluster_data_df = pd.DataFrame(Cluster_data)
        Cluster_data_df.to_excel(writer, index=False, sheet_name='Cluster')
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        sheet_name3 = 'Cluster'  
      # Access the sheet
        sheet = workbook[sheet_name3]
       # Remove the first row
        sheet.delete_cols(4)

        sheet.delete_cols(4)
        
        new_df.to_excel(writer, index=False, sheet_name='Losses and Delays')
        
        
        
        workbook = writer.book
        worksheet = workbook['Cluster'] 
        set_first_row_style(worksheet)
        
        
        
        Losses_and_Delays_df = pd.DataFrame(Losses_and_Delays)
        Losses_and_Delays_df.to_excel(writer, index=False, sheet_name='Losses and Delays')
        
       
        
       
        new_df.to_excel(writer, index=False, sheet_name='eUtran NeighRelations co-sites')
        
        workbook = writer.book
        worksheet = workbook['Losses and Delays']
        set_first_row_style(worksheet)
        
        
        
        eUtran_NeighRelations_co_sites_df = pd.DataFrame(eUtran_NeighRelations_co_sites)
        eUtran_NeighRelations_co_sites_df.to_excel(writer, index=False, sheet_name='eUtran NeighRelations co-sites')
        
        
        
        
        
        
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        new_df.to_excel(writer, index=False, sheet_name='LTE-UMTS (UtranFreqRelation)')
        
        workbook = writer.book
        worksheet = workbook['eUtran NeighRelations co-sites']
        set_first_row_style(worksheet)
        
        
        
        LTE_UMTS_UtranFreqRelation_df = pd.DataFrame(LTE_UMTS_UtranFreqRelation)
        LTE_UMTS_UtranFreqRelation_df.to_excel(writer, index=False, sheet_name='LTE-UMTS (UtranFreqRelation)')
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        new_df.to_excel(writer, index=False, sheet_name='LTE-LTE (EUtranFreqRelation)')
        
        workbook = writer.book
        worksheet = workbook['LTE-UMTS (UtranFreqRelation)']
        set_first_row_style(worksheet)
        
        
        
        LTE_LTE_EUtranFreqRelation_df = pd.DataFrame(LTE_LTE_EUtranFreqRelation)
        LTE_LTE_EUtranFreqRelation_df.to_excel(writer, index=False, sheet_name='LTE-LTE (EUtranFreqRelation)')
        column_width=15
        for sheet in workbook:
        # Adjust the width for all columns in the sheet
         for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    # Get the length of the cell value and set the width accordingly
                    cell_length = len(str(cell.value))
                    adjusted_width = (cell_length + 2) if cell_length >= 10 else column_width
                    sheet.column_dimensions[cell.column_letter].width = adjusted_width
        
        # new_df.to_excel(writer, index=False, sheet_name='Losses and Delays')
        
        workbook = writer.book
        worksheet = workbook['LTE-LTE (EUtranFreqRelation)']
        set_first_row_style(worksheet)

      
       

        # Save the Excel file using the workbook's save method
     
        workbook.save(new_file_path)
        
        


    except Exception as e:
        print("Error occurred:", e)

# Example usage:
# if len(sys.argv) != 2:
#         print("Usage: python final_tool.py <input_file>")
#         sys.exit(1)

input_file = 'media/output.xlsx'
search_variables = ["MARKET CLUSTER:","4-9 DIGIT SITE ID:","Approved? (Y/N):","REGION:", "LAT (DEC. DEG.):","CELL ID / BCF:", "LONG (DEC. DEG.):","ADDRESS:","STRUCTURE TYPE:","CTS Common ID","CSS - SECONDARY FUNCTION ID:", "AZIMUTH", "COUNTY:", "Soft Sector IDs", "ANTENNA MAKE - MODEL", "RRH - WCS band (QTY/MODEL)"]
search_excel_for_variables(input_file, search_variables)


import re

    
def update_eutran_parameters(input_file):
    input_sheet = "Sheet5"  # Assuming you want to use the same sheet as mentioned in the example

    # Find the row and column of the "Soft Sector IDs" cell in the input_file
    start_row_str, start_column_values, output_values = find_and_extract_data(input_file, input_sheet)

    if start_row_str is None or start_column_values is None:
        print("Data not found.")
        return

    # Convert start_row to an integer
    start_row = int(start_row_str)
    column_index = int(start_column_values)
    output_file ='media/Pushpendra_CIQ.xlsx'


    # Load the Pushpendra_CIQ.xlsx file
    wb = openpyxl.load_workbook(output_file)

    # Select the "eUtran Parameters" sheet
    output_sheet_name = "eUtran Parameters"
    output_sheet = wb[output_sheet_name]
    
  

    # Start from the specified row and column and put each value from the extracted data on a new row
    row_index = 2  # Start from row 2 in the output sheet
    for value in output_values:  # Loop through the extracted values
        output_sheet.cell(row=row_index, column=column_index, value=value)
        row_index += 1
        
        
    output_sheet_name = "Losses and Delays"
    output_sheet = wb[output_sheet_name]   
    row_index = 2
    col_index = 2
    # Start from row 2 in the output sheet
    for value in output_values:  # Loop through the extracted values
        output_sheet.cell(row=row_index, column=1, value=value)
        row_index += 1
        col_index += 0
        
    output_sheet_name = "PCI"
    output_sheet = wb[output_sheet_name]   
    row_index = 2
    col_index = 1
    # Start from row 2 in the output sheet
    for value in output_values:  # Loop through the extracted values
        output_sheet.cell(row=row_index, column=1, value=value)
        row_index += 1
        col_index += 0
        
    output_sheet_name = "eUtran NeighRelations co-sites"   
    output_sheet = wb[output_sheet_name] 
    row_index = 2
    col_index = 1
    value_mapping = {}  # To keep track of values in column 4
    for value in output_values:  # Loop through the extracted values
        output_sheet.cell(row=row_index, column=1, value=value)

        # Logic to determine the value for column 4 based on the value in column 1
        prefix = value.split("_")[1]  # Extract the prefix (e.g., 3A, 7A, 2A, etc.)
        base_value = value.replace(f"_{prefix}_", "_")  # Get the base value without the prefix

        # Reconstruct the new value for column 4 based on the pattern you provided
        if prefix.endswith("A"):
            new_prefix = prefix.replace("A", "B")
        elif prefix.endswith("B"):
            new_prefix = prefix.replace("B", "A")
        elif prefix.endswith("C"):
            new_prefix = prefix.replace("C", "A")  
          
        else:
            new_prefix = prefix  # Keep the same prefix if it's not A or B

        index = 1
        new_value = f"{base_value}{new_prefix}_{index}"

        while new_value in output_values or new_value in value_mapping:
         index += 1
         new_value = f"{base_value}{new_prefix}_{index}"

         value_mapping[new_value] = True

        output_sheet.cell(row=row_index, column=4, value=new_value)
        row_index += 1
        col_index += 0
        
        
    row_index = 2
    col_index = 1
    # Start from row 2 in the output sheet
    for value in output_values:  # Loop through the extracted values
        output_sheet.cell(row=row_index, column=1, value=value)
        row_index += 1
        col_index += 0       

    # Save the changes to the same file, replacing the old one
    wb.save(output_file)

def find_and_extract_data(input_file, start_sheet):
    
    
    
    wb = openpyxl.load_workbook(input_file)

    # Initialize variables to keep track of the latest occurrence of 'Soft Sector IDs'
    latest_cell = None
    latest_cell_row = 0
    target_sheet = wb[start_sheet]

    # Start the search from the specified row and column
    for row in target_sheet.iter_rows():
        for cell in row:
            if cell.value == 'Soft Sector IDs':
                latest_cell = cell
                latest_cell_row = cell.row

    # If 'Soft Sector IDs' is found, move to the cell to its immediate right
    if latest_cell:
        target_column = latest_cell.column + 1
        output_values = []
        while target_sheet.cell(row=latest_cell_row, column=target_column).value is not None:
            output_values.append(target_sheet.cell(row=latest_cell_row, column=target_column).value)
            latest_cell_row += 1

        return latest_cell_row, target_column, output_values
    
    
    
      
    # If 'Soft Sector IDs' is not found, return None
    return None, None, None

        

# Example usage:
# input_file = "file2.xlsx"
update_eutran_parameters(input_file)
sheet_name = 'Feeder Misc'  # Replace with the sheet name you want to edit
output_file = 'media/Pushpendra_CIQ.xlsx'
def edit_excel_file(output_file, sheet_name):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(output_file)

        # Access the sheet
        sheet = wb[sheet_name]

        # Remove the first row
        sheet.delete_rows(1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                cell.fill = PatternFill(start_color=None, end_color=None, fill_type='none')

        # Change the color of the new first row (previously the second row)
        for row in sheet.iter_rows(min_row=1, max_row=1 , max_col=16):
            for cell in row:
                cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                # Set the font color to white
                cell.font = Font(color='FFFFFF')
                cell.alignment = Alignment(vertical='center', horizontal='center')


                merge_ranges = [(4, 5), (6, 7), (8, 9), (10, 11), (12, 13), (14, 15)]  # (start_column, end_column)
                values = ['700 Band', '850 Band', '1700 Band', '1900 Band', '2100 Band', 'Delay']  # Values for the merged cells
                for i, (start_col, end_col) in enumerate(merge_ranges):
                    start_cell = sheet.cell(row=2, column=start_col)
                    end_cell = sheet.cell(row=2, column=end_col)
                    sheet.merge_cells(start_row=2, start_column=start_cell.column, end_row=2, end_column=end_cell.column)
                    start_cell.value = values[i]





        for row in sheet.iter_rows(min_row=2, max_row=3,max_col=16):
            for cell in row:
                cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                # Set the font color to white
                cell.font = Font(color='FFFFFF')
        
        sheet.sheet_properties.tabColor = "FFFFFFFF"

        # Save the changes to the workbook
        wb.save(output_file)
        wb.close()

      
    except Exception as e:
        print(f"An error occurred: {e}")



sheet_name2 = 'Revision History'  # Replace with the sheet name you want to edit

def edit_excel_file2(output_file, sheet_name2):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(output_file)

        # Access the sheet
        sheet = wb[sheet_name2]

        # Remove the first row
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        for row in sheet.iter_rows(min_row=1, max_row=4):
            for cell in row:
                cell.fill = PatternFill(start_color=None, end_color=None, fill_type='none')

        # Change the color of the new first row (previously the second row)
        for row in sheet.iter_rows(min_row=1, max_row=1 , min_col=1 , max_col=4):
            for cell in row:
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                # Set the font color to white
                cell.font = Font(color='000000')
                cell.alignment = Alignment(vertical='center', horizontal='center')


                merge_ranges = [(2, 4)]  # (start_column, end_column)
                values = ['Column name/position changes, cosmetic changes - version number + 0.1 Adding columns/tabs - version number + 1']  # Values for the merged cells
                for i, (start_col, end_col) in enumerate(merge_ranges):
                    start_cell = sheet.cell(row=2, column=start_col)
                    end_cell = sheet.cell(row=2, column=end_col)
                    sheet.merge_cells(start_row=2, start_column=start_cell.column, end_row=2, end_column=end_cell.column)
                    start_cell.value = values[i]


        for row in sheet.iter_rows(min_row=2, max_row=3,max_col=4):
            for cell in row:
                cell.fill = PatternFill(start_color='73DCFF', end_color='73DCFF', fill_type='solid')
                # Set the font color to white
                cell.font = Font(color='000000')
        


        for row in sheet.iter_rows(min_row=4, max_row=4,max_col=4):
            for cell in row:
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                # Set the font color to white
                cell.font = Font(color='000000')
                
                column_widths = [25, 88, 30, 34]
                for col, width in enumerate(column_widths, start=1):
                    sheet.column_dimensions[sheet.cell(row=4, column=col).column_letter].width = width 
        

        # Save the changes to the workbook
        wb.save(output_file)
        wb.close()

       
    except Exception as e:
        print(f"An error occurred: {e}")
        
        
        
# Call the function to edit the Excel file
sheet_name4 = 'Losses and Delays'
cell_width=10
cell_height=150
def set_vertical_text_in_cells(output_file, sheet_name4 , cell_width, cell_height):
    # Load the existing Excel file or create a new one
    workbook = openpyxl.load_workbook(output_file)
    
    # Select the desired sheet
    sheet = workbook[sheet_name4]

    # Set the column width and row height
    for column in sheet.columns:
        col_letter = column[0].column_letter
        sheet.column_dimensions[col_letter].width = cell_width

    sheet.row_dimensions[1].height = cell_height  # Set the height of the first row

    # Apply text rotation to each cell in the first row
    first_row = sheet[1]
    for cell in first_row:
        if cell.value:
            cell.alignment = Alignment(textRotation=90 , wrapText=True)
            cell.value = cell.value  
            
    column_widths = [18]
    for col, width in enumerate(column_widths, start=1):
                    sheet.column_dimensions[sheet.cell(row=4, column=col).column_letter].width = width         

    # Save the modified Excel file
    workbook.save(output_file)
    


columns_to_lock_per_sheet = {
    3: [1, 2],  # Lock columns 1 and 2 in sheet 2
    4: [1],     # Lock column 3 in sheet 5
    7: [1],     # Lock column 5 in sheet 8
    8: [1],     # Lock column 5 in sheet 8
    9: [1],     # Lock column 5 in sheet 8
}
def lock_columns(output_file, columns_to_lock_per_sheet):
    # Load the Excel file
    wb = openpyxl.load_workbook(output_file)
    for sheet_num, columns_to_lock in columns_to_lock_per_sheet.items():
       # Get the specified sheet
        sheet_name = wb.sheetnames[sheet_num - 1]
        sheet = wb[sheet_name]
        # Calculate the column index where we want to freeze the panes
        freeze_column_index = max(columns_to_lock) + 1 if columns_to_lock else 1
        # Freeze the panes to make specified columns fixed
        sheet.freeze_panes = sheet.cell(row=1, column=freeze_column_index)
    # Save the modified Excel fil
    wb.save(output_file)   
    




import re

def find_value_between_underscores(input_data):
    # Check if "F" is present in the input_data
     # Check if "F" is present in the input_data
    # Check if "F" is present in the input_data
    if "F" in input_data:
        # Extract the part between the first two underscores
        pattern_f = re.compile(r'^([A-Z0-9]+)_([^_]+)_([0-9]+)_F.*$')
        match_f = pattern_f.search(input_data)

        if match_f:
            prefix, value_between_underscores, number = match_f.groups()
            alphabets = ['A', 'B', 'C']

            if value_between_underscores[-1] in alphabets:
                remaining_alphabets = [letter for letter in alphabets if letter != value_between_underscores[-1]]
                new_values = [prefix + "_" + value_between_underscores[:-1] + letter + "_" + number + "_F" for letter in remaining_alphabets]
                return " ".join(new_values)
            else:
                return None
        else:
            return None
    else:
        pattern = re.compile(r'^[A-Z0-9]+_([^_]+)_.*$')
        match = pattern.search(input_data)

        if match:
            value_between_underscores = match.group(1)
            prefix = input_data.split("_")[0]  # Extract the prefix before the first underscore
            alphabet = value_between_underscores[-1]  # Extract the last alphabet from the value between underscores
            number = input_data.split("_")[-1]  # Extract the number after the last underscore
            alphabets = ['A', 'B','C']

            if alphabet in alphabets:
                remaining_alphabets = [letter for letter in alphabets if letter != alphabet]
                new_values = [prefix + "_" + value_between_underscores[:-1] + letter + "_" + number for letter in remaining_alphabets]
                return " ".join(new_values)
            else:
                return None
        else:
            return None

# Load the Excel file
excel_file_path ='media/Pushpendra_CIQ.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook["eUtran NeighRelations co-sites"]  # Assuming the sheet name is "Sheet8"

# Get the maximum row number (number of rows in the sheet)
max_row = sheet.max_row

# Start from row 2 (assuming values are present from row 2 onwards in column 1)
start_row = 2

# Loop through each row in column 1, starting from row 2
for row in range(start_row, max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value
    if cell_value:
        # Apply the logic to find modified values using the function find_value_between_underscores()
        result = find_value_between_underscores(cell_value)
        if result:
            modified_values = result.split()
            # Write the modified values to the next three columns (B, C, and D)
            for idx, value in enumerate(modified_values):
                sheet.cell(row=row, column=2 + idx, value=value)

# Save the changes to the Excel file
    result = find_value_between_underscores(cell_value)
   

excel_file_path ='media/Pushpendra_CIQ.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook["eUtran NeighRelations co-sites"]  # Assuming the sheet name is "Sheet8"

# Get the maximum row number (number of rows in the sheet)
max_row = sheet.max_row

# Start from row 2 (assuming values are present from row 2 onwards in column 1)
start_row = 2

# Loop through each row in column 1, starting from row 2
for row in range(start_row, max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value
    if cell_value:
        # Apply the logic to find modified values using the function find_value_between_underscores()
        result = find_value_between_underscores(cell_value)
        if result:
            modified_values = result.split()
            if len(modified_values) == 2:
                # Write the modified values to columns 5 and 6 (E and F)
                sheet.cell(row=row, column=4, value=modified_values[0])
                sheet.cell(row=row, column=5, value=modified_values[1])

# Save the changes to the Excel file
workbook.save(excel_file_path)







edit_excel_file(output_file, sheet_name)
edit_excel_file2(output_file, sheet_name2)
set_vertical_text_in_cells(output_file, sheet_name4, cell_width, cell_height)
lock_columns(output_file, columns_to_lock_per_sheet)
print("success")
